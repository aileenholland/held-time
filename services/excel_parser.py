"""
excel_parser.py
---------------
Reads the weekly Billing v2 Excel export and the 2026 Held Time Tracker
from the shared reports folder, returning project-level fee and held-time data
keyed by project number (e.g. "1264").

Data sources
------------
1. Latest weekly export  (e.g. 03-16.xlsx)  — Sheet "Billing v2"
   Columns: Title | Status | SQFT | Design Fees Sold | Actual fees | Sold vs Spent | Lead Designer
   → provides: actual_fees, sold_vs_spent (and fills sqft / designer when Wrike is blank)

2. 2026 Held time tracker.xlsx  — Sheet "MASTER TRACKING"
   Columns: (blank) | Title | Status | SQFT | Fees Sold | Actual Fees | Sold vs Spent |
            Senior Designer | COs | Fee+CO | % phase | % fee | Held Time | HT Category |
            (blank) | Notes
   → provides: cos, fee_plus_co, pct_phase, pct_fee, held_time, ht_category, notes

3. 2026 Held time tracker.xlsx  — Sheet "Completed Projects"
   Row 1 = summary title, Row 2 = headers, data from Row 3
   Columns: (blank) | Title | Date | SQFT | Fees Sold | Actual Fees | Sold vs Spent |
            Designer | COs | Fee+CO | Profit/HT | POD | HT inc Feas | HT Reason
   → provides: date_completed, actual_fees, sold_vs_spent, cos, fee_plus_co,
                held_time, pod, notes
"""

import os
import re
import glob
import json
from datetime import datetime

import shutil
import tempfile
from openpyxl import load_workbook

# ── Overrides store (dashboard-editable fields: cos, notes) ─────────────────
_OVERRIDES_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'data', 'overrides.json'
)


def load_overrides():
    """Return dict of {project_number: {cos, notes}} dashboard overrides."""
    try:
        with open(_OVERRIDES_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_override(project_number, field, value):
    """Persist a single field override for a project. Returns updated overrides."""
    overrides = load_overrides()
    if project_number not in overrides:
        overrides[project_number] = {}
    overrides[project_number][field] = value
    os.makedirs(os.path.dirname(_OVERRIDES_PATH), exist_ok=True)
    with open(_OVERRIDES_PATH, 'w', encoding='utf-8') as f:
        json.dump(overrides, f, indent=2)
    return overrides

# ── Folder configuration ─────────────────────────────────────────────────────
REPORTS_FOLDER = os.environ.get(
    'REPORTS_FOLDER',
    r'C:\Users\Aileen Holland\Clearspace Offices\Design - Admin\Design Fees\Held time reports\2026',
)
TRACKER_FILENAME = '2026 Held time tracker.xlsx'


# ── Helpers ──────────────────────────────────────────────────────────────────

def _open_workbook(path):
    """Open a workbook, copying to a temp file first if it is locked by Excel."""
    try:
        return load_workbook(path, data_only=True)
    except PermissionError:
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        shutil.copy2(path, tmp.name)
        return load_workbook(tmp.name, data_only=True)


def _safe_float(val):
    """Return float or None; handles empty strings and Excel error strings."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '#VALUE!', '#REF!', '#N/A', '#DIV/0!', '#NULL!', '#NUM!'):
        return None
    try:
        return float(s.replace(',', ''))
    except (ValueError, TypeError):
        return None


def _extract_num(title):
    """Extract project number from '[1264] Some Name' → '1264'."""
    if not title:
        return None
    m = re.match(r'\[(\d+)\]', str(title).strip())
    return m.group(1) if m else None


# ── Public API ────────────────────────────────────────────────────────────────

def get_latest_billing_export():
    """
    Scan REPORTS_FOLDER for the most-recently-modified weekly export
    (any .xlsx that is not the tracker file).

    Returns
    -------
    filename : str | None
        Basename of the file used, or None if no file found.
    file_date : str | None
        Last-modified date as 'YYYY-MM-DD', or None.
    data : dict
        Mapping project_number → {actual_fees, sold_vs_spent,
                                   fees_sold_xl, sqft_xl, designer_xl, status_xl}
    """
    pattern = os.path.join(REPORTS_FOLDER, '*.xlsx')
    files = [
        f for f in glob.glob(pattern)
        if os.path.basename(f).lower() != TRACKER_FILENAME.lower()
    ]
    if not files:
        return None, None, {}

    latest = max(files, key=os.path.getmtime)
    filename = os.path.basename(latest)
    file_date = datetime.fromtimestamp(os.path.getmtime(latest)).strftime('%Y-%m-%d')

    wb = _open_workbook(latest)
    ws = wb['Billing v2'] if 'Billing v2' in wb.sheetnames else wb.active

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 6:
            continue
        title        = row[0]
        status       = row[1]
        sqft         = row[2]
        fees_sold    = row[3]
        actual_fees  = row[4]
        sold_vs_spent = row[5]
        designer     = row[6] if len(row) > 6 else None

        num = _extract_num(title)
        if not num:
            continue

        data[num] = {
            'actual_fees':    _safe_float(actual_fees),
            'sold_vs_spent':  _safe_float(sold_vs_spent),
            'fees_sold_xl':   _safe_float(fees_sold),
            'sqft_xl':        int(float(sqft)) if _safe_float(sqft) is not None else None,
            'designer_xl':    str(designer).strip() if designer else '',
            'status_xl':      str(status).strip() if status else '',
        }

    wb.close()
    return filename, file_date, data


def get_tracker_data():
    """
    Parse the 2026 Held time tracker workbook.

    Returns
    -------
    active_data : dict
        project_number → {cos, fee_plus_co, pct_phase, pct_fee,
                           held_time, ht_category, notes}
    completed_data : dict
        project_number → {date_completed, actual_fees, sold_vs_spent,
                           cos, fee_plus_co, held_time, pod, notes}
    """
    tracker_path = os.path.join(REPORTS_FOLDER, TRACKER_FILENAME)
    if not os.path.exists(tracker_path):
        return {}, {}

    wb = _open_workbook(tracker_path)

    # ── MASTER TRACKING ──────────────────────────────────────────────────────
    # 0-indexed cols:
    # A=0(blank) B=1(title) C=2(status) D=3(sqft) E=4(fees_sold)
    # F=5(actual_fees) G=6(sold_vs_spent) H=7(designer) I=8(cos)
    # J=9(fee+co) K=10(%phase) L=11(%fee) M=12(held_time)
    # N=13(ht_category) O=14(blank) P=15(notes)
    active_data = {}
    if 'MASTER TRACKING' in wb.sheetnames:
        ws = wb['MASTER TRACKING']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 13:
                continue
            num = _extract_num(row[1])
            if not num:
                continue
            notes = row[15] if len(row) > 15 else None
            active_data[num] = {
                'cos':         _safe_float(row[8]),
                'fee_plus_co': _safe_float(row[9]),
                'pct_phase':   _safe_float(row[10]),
                'pct_fee':     _safe_float(row[11]),
                'held_time':   _safe_float(row[12]),
                'ht_category': str(row[13]).strip() if row[13] else '',
                'notes':       str(notes).strip() if notes else '',
            }

    # ── Completed Projects ────────────────────────────────────────────────────
    # Row 1 = summary title, Row 2 = column headers, data from Row 3
    # 0-indexed cols:
    # A=0(blank) B=1(title) C=2(date) D=3(sqft) E=4(fees_sold)
    # F=5(actual_fees) G=6(sold_vs_spent) H=7(designer) I=8(cos)
    # J=9(fee+co) K=10(profit/HT) L=11(POD) M=12(HT inc Feas) N=13(HT reason)
    completed_data = {}
    if 'Completed Projects' in wb.sheetnames:
        ws = wb['Completed Projects']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 10:
                continue
            title = row[1]
            num = _extract_num(title)
            if not num:
                continue

            date_val = row[2]
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            elif date_val:
                date_str = str(date_val)[:10]
            else:
                date_str = None

            clean_name = re.sub(r'^\[\d+\]\s*', '', str(title).strip())
            sqft_raw   = row[3]
            completed_data[num] = {
                'name':           clean_name,
                'date_completed': date_str,
                'sqft':           int(float(sqft_raw)) if _safe_float(sqft_raw) is not None else None,
                'fees_sold':      _safe_float(row[4]),
                'actual_fees':    _safe_float(row[5]),
                'sold_vs_spent':  _safe_float(row[6]),
                'designer':       str(row[7]).strip() if len(row) > 7 and row[7] else '',
                'cos':            _safe_float(row[8]),
                'fee_plus_co':    _safe_float(row[9]),
                'held_time':      _safe_float(row[10]),
                'pod':            str(row[11]).strip() if len(row) > 11 and row[11] else '',
                'notes':          str(row[13]).strip() if len(row) > 13 and row[13] else '',
            }

    wb.close()
    return active_data, completed_data


def get_file_info():
    """
    Return metadata about the data files currently loaded.

    Returns
    -------
    dict with keys: billing_file, billing_date, tracker_exists
    """
    filename, file_date, _ = get_latest_billing_export()
    tracker_path = os.path.join(REPORTS_FOLDER, TRACKER_FILENAME)
    return {
        'billing_file':   filename,
        'billing_date':   file_date,
        'tracker_exists': os.path.exists(tracker_path),
    }
